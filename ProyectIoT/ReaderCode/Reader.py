import json
import git
import os
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Definición de la ruta del directorio del repositorio y la URL del repositorio en GitHub
REPO_DIR = os.path.join(os.getcwd(), 'Codigos')
REPO_URL = 'https://github.com/UlisesFox/Codigos.git'
JSON_FILE_PATH = os.path.join(REPO_DIR, 'codes.json')
EXCEL_FILE_PATH = os.path.join(REPO_DIR, 'Codes.xlsx')

# Clona el repositorio si no existe localmente, de lo contrario, abre el repositorio existente
if not os.path.exists(REPO_DIR):
    repo = git.Repo.clone_from(REPO_URL, REPO_DIR)
else:
    repo = git.Repo(REPO_DIR)

# Función para actualizar el repositorio local con los cambios del remoto
def update_repo():
    origin = repo.remote(name='origin')
    origin.pull()

# Función para leer el archivo JSON y devolver su contenido
def read_json_file():
    if os.path.exists(JSON_FILE_PATH):
        with open(JSON_FILE_PATH, 'r') as file:
            data = json.load(file)
            return data
    else:
        print("No se encontró el archivo JSON.")
        return []

# Función para guardar los datos en un archivo Excel
def save_to_excel(data):
    if os.path.exists(EXCEL_FILE_PATH):
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['codigo', 'Fecha completa', 'Cantidad'])

    # Obtiene la fecha actual
    current_time_all = time.strftime('%Y-%m-%d')

    # Actualiza o agrega las filas en el archivo Excel
    for code in data:
        found = False
        for row in sheet.iter_rows(min_row=2):
            if code == row[0].value and current_time_all == row[1].value:
                row[2].value += 1
                found = True
                break
        
        if not found:
            can = 1
            sheet.append([code, current_time_all, can])

    # Guarda los cambios en el archivo Excel
    workbook.save(EXCEL_FILE_PATH)

# Bucle principal
while True:
    update_repo()  # Actualiza el repositorio local
    data = read_json_file()  # Lee los datos del archivo JSON
    save_to_excel(data)  # Guarda los datos en el archivo Excel
    print('Datos actualizados! (Cooldown 15s)')
    time.sleep(15)  # Espera 15 segundos antes de repetir el proceso
